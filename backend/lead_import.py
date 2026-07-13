"""Lead import — parses xlsx/csv/Google-Sheet leads files into Account rows.

Two-step flow, no surprise writes:
  preview(content, filename, db) -> parses + maps columns + flags dupes. No DB writes.
  commit(rows, filename, db)     -> creates/updates Account rows from previewed rows.

Never synthesizes a missing email address — many real lead sheets (e.g. FaithForge's
Top-25 Maryland/DC sheet) intentionally leave email blank with "needs research" notes,
and guessing first.last@company.com is explicitly forbidden.
"""
import csv
import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import requests
from sqlalchemy.orm import Session

from models import Account, ACCOUNT_STAGES

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Fields every parsed lead dict always carries (defaults to "").
_LEAD_FIELDS = (
    "company_name", "segment", "website", "location", "contact_name", "contact_title",
    "contact_email", "contact_phone", "stage", "pain_points", "entry_offer",
)


# ── Header normalization + mapping ──────────────────────────────────────────

def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[\/_\-]+", " ", s)
    s = re.sub(r"[^a-z0-9 ]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# Exact-match aliases (normalized header -> target field). Reference-sheet headers
# (FaithForge_Top_25_Maryland_DC_Lead_Sheet.xlsx) are listed first since they're the
# canonical format; generic aliases follow so other spreadsheets still import.
_ALIASES: Dict[str, str] = {
    # Reference sheet — canonical format
    "priority": "priority_rank",
    "name": "contact_name",
    "title": "contact_title",
    "company org": "company_name",
    "location": "location",
    "fit for faithforge": "fit_notes",
    "targeted gap angle": "pain_points",
    "how bernedette can help": "entry_offer",
    "linkedin profile": "linkedin_url",
    "source url": "source_url",
    "email status": "email_status",
    "email contact notes": "contact_notes",
    "first email linkedin message": "sample_message",  # this is the OUTPUT we generate — ignore on import
    "follow up status": "stage",
    "next step": "next_action",
    # Generic aliases
    "company": "company_name",
    "organization": "company_name",
    "org": "company_name",
    "account": "company_name",
    "company name": "company_name",
    "contact": "contact_name",
    "contact name": "contact_name",
    "full name": "contact_name",
    "email": "contact_email",
    "e mail": "contact_email",
    "email address": "contact_email",
    "contact email": "contact_email",
    "job title": "contact_title",
    "role": "contact_title",
    "position": "contact_title",
    "segment": "segment",
    "industry": "segment",
    "sector": "segment",
    "city": "location",
    "state": "location",
    "region": "location",
    "phone": "contact_phone",
    "mobile": "contact_phone",
    "tel": "contact_phone",
    "telephone": "contact_phone",
    "pain": "pain_points",
    "pain points": "pain_points",
    "entry offer": "entry_offer",
    "website": "website",
    "url": "website",
    "domain": "website",
    "notes": "notes",
    "status": "stage",
}

_ALIASES_BY_LENGTH = sorted(_ALIASES, key=len, reverse=True)


def _map_header(header: str) -> Optional[str]:
    key = _norm(header)
    if key in _ALIASES:
        return _ALIASES[key]
    for alias in _ALIASES_BY_LENGTH:
        if alias and alias in key:
            return _ALIASES[alias]
    return None


def auto_map_columns(headers: List[str]) -> Dict[str, Optional[str]]:
    """header -> mapped field, or None if unrecognized."""
    return {h: _map_header(h) for h in headers}


# ── File parsing ─────────────────────────────────────────────────────────────

def parse_xlsx(content: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, [])
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows = []
    for raw in rows_iter:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = raw[i] if i < len(raw) else None
            row[h] = "" if val is None else str(val).strip()
        rows.append(row)
    return headers, rows


def parse_csv(content: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = []
    for raw in reader:
        if not any((v or "").strip() for v in raw.values()):
            continue
        rows.append({k: (v or "").strip() for k, v in raw.items() if k})
    return headers, rows


def google_sheet_csv_url(share_url: str) -> str:
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", share_url)
    if not m:
        raise ValueError("That doesn't look like a Google Sheets link.")
    sheet_id = m.group(1)
    gid_match = re.search(r"[?&#]gid=(\d+)", share_url)
    gid = gid_match.group(1) if gid_match else "0"
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def fetch_google_sheet(share_url: str) -> bytes:
    url = google_sheet_csv_url(share_url)
    resp = requests.get(url, timeout=30)
    if resp.status_code != 200 or resp.headers.get("content-type", "").startswith("text/html"):
        raise ValueError(
            "Could not fetch that Google Sheet. Make sure it's shared as "
            "'Anyone with the link can view', or export it to XLSX/CSV and upload directly."
        )
    return resp.content


# ── Row -> lead normalization ────────────────────────────────────────────────

def _normalize_stage(value: str) -> str:
    v = (value or "").strip().lower()
    for stage in ACCOUNT_STAGES:
        if stage.lower() == v:
            return stage
    return "Not Contacted"


def _build_lead(row: Dict[str, str], mapping: Dict[str, Optional[str]]) -> Dict[str, Any]:
    lead: Dict[str, Any] = {f: "" for f in _LEAD_FIELDS}
    lead["priority_rank"] = None
    note_parts: List[str] = []
    linkedin_url = ""

    for header, value in row.items():
        if not value:
            continue
        field = mapping.get(header)
        if field is None:
            continue

        if field == "priority_rank":
            digits = re.sub(r"[^\d]", "", value)
            if digits:
                lead["priority_rank"] = int(digits)
        elif field == "linkedin_url":
            linkedin_url = value
        elif field == "fit_notes":
            note_parts.append(f"Fit for FaithForge: {value}")
        elif field == "source_url":
            note_parts.append(f"Source: {value}")
        elif field == "email_status":
            note_parts.append(f"Email status: {value}")
        elif field == "contact_notes":
            note_parts.append(f"Contact notes: {value}")
        elif field == "sample_message":
            continue  # output column, not an input
        elif field == "stage":
            lead["stage"] = _normalize_stage(value)
        elif field == "contact_email":
            candidate = value.strip()
            if EMAIL_RE.match(candidate):
                lead["contact_email"] = candidate
            else:
                note_parts.append(f"Unrecognized email value ignored: {candidate}")
        elif field in lead:
            lead[field] = value
        elif field == "notes":
            note_parts.append(value)

    if linkedin_url:
        if not lead.get("website"):
            lead["website"] = linkedin_url
        else:
            note_parts.append(f"LinkedIn: {linkedin_url}")

    lead["notes"] = "\n".join(note_parts)
    lead["has_email"] = bool(lead.get("contact_email"))
    return lead


def _apply_priority_scores(rows: List[Dict[str, Any]]) -> None:
    """Convert this upload's rank column (1 = best) into a 0-100 priority_score,
    scaled to the ranks actually present in this file."""
    ranked = [r for r in rows if r.get("priority_rank")]
    if not ranked:
        return
    max_rank = max(r["priority_rank"] for r in ranked)
    span = max(max_rank - 1, 1)
    for r in ranked:
        r["priority_score"] = round(100 - (r["priority_rank"] - 1) * 100 / span, 1)


# ── Duplicate detection ──────────────────────────────────────────────────────

def _find_duplicate(db: Session, lead: Dict[str, Any]) -> Optional[Account]:
    email = (lead.get("contact_email") or "").strip().lower()
    if email:
        return db.query(Account).filter(Account.contact_email.ilike(email)).first()
    company = (lead.get("company_name") or "").strip().lower()
    name = (lead.get("contact_name") or "").strip().lower()
    if not company:
        return None
    q = db.query(Account).filter(Account.company_name.ilike(company))
    if name:
        q = q.filter(Account.contact_name.ilike(name))
    return q.first()


# ── Preview / commit ─────────────────────────────────────────────────────────

def preview(content: bytes, filename: str, db: Session) -> Dict[str, Any]:
    ext = (filename or "").rsplit(".", 1)[-1].lower()
    if ext in ("xlsx", "xls"):
        headers, raw_rows = parse_xlsx(content)
    elif ext == "csv":
        headers, raw_rows = parse_csv(content)
    else:
        raise ValueError("Unsupported file type. Upload .xlsx, .xls, or .csv.")

    if not headers or not raw_rows:
        raise ValueError("No rows found in that file.")

    mapping = auto_map_columns(headers)
    rows = []
    for raw in raw_rows:
        lead = _build_lead(raw, mapping)
        if not lead.get("company_name") and not lead.get("contact_name"):
            continue  # skip fully blank/unusable rows
        rows.append(lead)

    _apply_priority_scores(rows)

    email_missing = 0
    duplicate_count = 0
    for lead in rows:
        dup = _find_duplicate(db, lead)
        lead["duplicate_of_account_id"] = dup.id if dup else None
        lead["duplicate_company"] = dup.company_name if dup else None
        if dup:
            duplicate_count += 1
        if not lead["has_email"]:
            email_missing += 1

    return {
        "columns": headers,
        "mapping": mapping,
        "rows": rows,
        "row_count": len(rows),
        "duplicate_count": duplicate_count,
        "email_missing_count": email_missing,
    }


def preview_google_sheet(share_url: str, db: Session) -> Dict[str, Any]:
    content = fetch_google_sheet(share_url)
    return preview(content, "google_sheet.csv", db)


def commit(rows: List[Dict[str, Any]], source_filename: str, db: Session,
           dedupe: str = "skip") -> Dict[str, Any]:
    """dedupe: 'skip' (leave existing Account untouched) | 'update' (merge new fields in)."""
    created, updated, skipped = 0, 0, 0
    account_ids: List[int] = []

    for lead in rows:
        if not lead.get("company_name") and not lead.get("contact_name"):
            skipped += 1
            continue
        existing = _find_duplicate(db, lead)
        if existing:
            if dedupe == "update":
                for field in ("segment", "website", "location", "contact_title", "contact_phone",
                              "pain_points", "entry_offer"):
                    val = lead.get(field)
                    if val and not getattr(existing, field, None):
                        setattr(existing, field, val)
                if lead.get("contact_email") and not existing.contact_email:
                    existing.contact_email = lead["contact_email"]
                if lead.get("priority_score") is not None and existing.priority_score is None:
                    existing.priority_score = lead["priority_score"]
                if lead.get("notes"):
                    existing.notes = f"{existing.notes}\n{lead['notes']}" if existing.notes else lead["notes"]
                existing.updated_at = datetime.utcnow()
                updated += 1
            else:
                skipped += 1
            account_ids.append(existing.id)
            continue

        acc = Account(
            company_name=lead.get("company_name") or lead.get("contact_name") or "Unknown",
            segment=lead.get("segment") or None,
            website=lead.get("website") or None,
            location=lead.get("location") or None,
            contact_name=lead.get("contact_name") or None,
            contact_title=lead.get("contact_title") or None,
            contact_email=lead.get("contact_email") or None,
            contact_phone=lead.get("contact_phone") or None,
            stage=lead.get("stage") or "Not Contacted",
            priority_score=lead.get("priority_score"),
            pain_points=lead.get("pain_points") or None,
            entry_offer=lead.get("entry_offer") or None,
            notes=lead.get("notes") or None,
            source=f"bulk_upload:{source_filename}",
        )
        db.add(acc)
        db.flush()  # assign acc.id without a full commit
        account_ids.append(acc.id)
        created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped, "account_ids": account_ids}
